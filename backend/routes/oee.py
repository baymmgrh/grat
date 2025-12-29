from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, Machine, MaintenanceRecord, MaintenanceSchedule, User
from utils.i18n import success_response, error_response, get_message
from models.oee import OEERecord, OEEDowntimeRecord, OEETarget, OEEAlert, MaintenanceImpact, OEEAnalytics, QualityDefect
from models.product_excel_schema import ProductNew
from utils import generate_number
from datetime import datetime, date, timedelta
from sqlalchemy import func, and_, or_, desc
import json
import re
import io

oee_bp = Blueprint('oee', __name__)

@oee_bp.route('/records', methods=['GET'])
@oee_bp.route('/records/', methods=['GET'])
@jwt_required()
def get_records():
    """Get OEE records from both OEERecord and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        machine_id = request.args.get('machine_id', type=int)
        limit = request.args.get('limit', 100, type=int)
        
        all_records = []
        
        # Get from OEERecord
        oee_query = OEERecord.query
        if machine_id:
            oee_query = oee_query.filter(OEERecord.machine_id == machine_id)
        
        oee_records = oee_query.order_by(OEERecord.record_date.desc()).limit(limit).all()
        for r in oee_records:
            all_records.append({
                'id': r.id,
                'source': 'oee_record',
                'record_number': r.record_number,
                'machine_id': r.machine_id,
                'machine_name': r.machine.name if r.machine else None,
                'record_date': r.record_date.isoformat() if r.record_date else None,
                'availability': float(r.availability) if r.availability else 0,
                'performance': float(r.performance) if r.performance else 0,
                'quality': float(r.quality) if r.quality else 0,
                'oee_percentage': float(r.oee) if r.oee else 0
            })
        
        # Get from ShiftProduction
        shift_query = ShiftProduction.query
        if machine_id:
            shift_query = shift_query.filter(ShiftProduction.machine_id == machine_id)
        
        shift_records = shift_query.order_by(ShiftProduction.production_date.desc()).limit(limit).all()
        for sp in shift_records:
            # Get product name from work order or product directly
            product_name = None
            target_qty = int(sp.target_quantity) if sp.target_quantity else 0
            actual_qty = int(sp.actual_quantity) if sp.actual_quantity else 0
            pack_per_karton = 50  # Default value
            
            # Try to get product name from product or work_order
            if sp.product:
                product_name = sp.product.name
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
            
            # Get pack_per_karton from ProductNew if product name available
            if product_name:
                try:
                    # Remove "WIP " prefix for matching
                    search_name = product_name.replace('WIP ', '').strip()
                    product_new = ProductNew.query.filter(
                        ProductNew.nama_produk.ilike(f'%{search_name}%')
                    ).first()
                    if product_new and product_new.pack_per_karton:
                        pack_per_karton = int(product_new.pack_per_karton)
                except Exception as e:
                    # If query fails, use default
                    pack_per_karton = 50
            
            # Get target from work order if not set
            if not target_qty and sp.work_order:
                target_qty = int(sp.work_order.quantity) if sp.work_order.quantity else 0
            
            # Parse ALL downtime entries from issues field
            # Format: "60 menit - Produk bocor (endseal kotor) [others]; 20 menit - Kain keluar jalur [others]; ..."
            downtime_breakdown = []
            total_downtime_minutes = 0  # Sum of NON-IDLE downtime entries only
            idle_time_minutes = 0  # Separate from downtime
            
            # IDLE TIME keywords - waiting for materials/resources
            idle_keywords = [
                'tunggu kain', 'tunggu stiker', 'tunggu packaging', 'tunggu mixing',
                'tunggu bahan', 'tunggu material', 'tunggu label', 'tunggu box',
                'tunggu karton', 'tunggu lem', 'tunggu tinta', 'tunggu order',
                'menunggu kain', 'menunggu stiker', 'menunggu packaging', 'menunggu mixing',
                'nunggu kain', 'nunggu stiker', 'nunggu packaging', 'nunggu mixing',
                'waiting for', 'standby'
            ]
            
            if sp.issues:
                # Split by semicolon
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    # Parse: "60 menit - Produk bocor (endseal kotor) [others]"
                    # More flexible regex - captures number and reason
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason = match.group(2).strip()
                        # Remove trailing category bracket if present [category]
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason).strip()
                        reason_lower = reason.lower()
                        
                        # Check if this is idle time (tunggu keywords)
                        is_idle = any(kw in reason_lower for kw in idle_keywords)
                        if is_idle:
                            # Idle time is SEPARATE from downtime
                            idle_time_minutes += duration
                        else:
                            # Only non-idle goes to downtime
                            total_downtime_minutes += duration
                        
                        # Add to breakdown for Top 3 display
                        downtime_breakdown.append({'reason': reason, 'duration_minutes': duration, 'is_idle': is_idle})
            
            # Fallback: use sum of category fields if issues parsing yielded 0
            if total_downtime_minutes == 0:
                # Try downtime_minutes field first
                if sp.downtime_minutes:
                    total_downtime_minutes = int(sp.downtime_minutes)
                else:
                    # Sum from individual category fields
                    total_downtime_minutes = (
                        (int(sp.downtime_mesin) if sp.downtime_mesin else 0) +
                        (int(sp.downtime_operator) if sp.downtime_operator else 0) +
                        (int(sp.downtime_material) if sp.downtime_material else 0) +
                        (int(sp.downtime_design) if sp.downtime_design else 0) +
                        (int(sp.downtime_others) if sp.downtime_others else 0)
                    )
            
            # Sort by duration descending and take top 3
            downtime_breakdown.sort(key=lambda x: x['duration_minutes'], reverse=True)
            top_3_downtime = downtime_breakdown[:3]
            
            # Extract shift number from shift field (e.g., "shift_1" -> 1)
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Get planned runtime per shift (Shift 1=510, Shift 2=480, Shift 3=450 minutes)
            if shift_num == 1:
                default_planned = 510  # 06:30-15:00 = 8.5 hours
            elif shift_num == 2:
                default_planned = 480  # 15:00-23:00 = 8 hours
            else:
                default_planned = 450  # 23:00-06:30 = 7.5 hours
            
            planned_runtime = int(sp.planned_runtime) if sp.planned_runtime else default_planned
            
            # Runtime = Planned - Downtime - Idle (calculated, not from input)
            runtime_minutes = max(0, planned_runtime - total_downtime_minutes - idle_time_minutes)
            
            # Shift data with Grade A, B, C and Runtime/Downtime/Idle metrics
            shift_data = {
                'shift': shift_num,
                'grade_a': int(sp.good_quantity) if sp.good_quantity else 0,
                'grade_b': int(sp.rework_quantity) if sp.rework_quantity else 0,
                'grade_c': int(sp.reject_quantity) if sp.reject_quantity else 0,
                'total': int(sp.actual_quantity) if sp.actual_quantity else 0,
                'runtime_minutes': runtime_minutes,
                'downtime_minutes': total_downtime_minutes,
                'idle_time_minutes': idle_time_minutes,
                'planned_runtime': planned_runtime
            }
            
            all_records.append({
                'id': sp.id,
                'source': 'shift_production',
                'record_number': f"SP-{sp.id}",
                'machine_id': sp.machine_id,
                'machine_name': sp.machine.name if sp.machine else None,
                'record_date': sp.production_date.isoformat() if sp.production_date else None,
                'availability': float(sp.efficiency_rate) if sp.efficiency_rate else 0,
                'performance': 100.0,  # Default performance
                'quality': float(sp.quality_rate) if sp.quality_rate else 0,
                'oee_percentage': float(sp.oee_score) if sp.oee_score else 0,
                'product_name': product_name,
                'target_quantity': target_qty,
                'actual_quantity': actual_qty,
                'top_3_downtime': top_3_downtime,
                'shift_data': shift_data,
                'pack_per_karton': pack_per_karton,
                # New fields for runtime/downtime tracking
                'runtime_minutes': runtime_minutes,
                'total_downtime_minutes': total_downtime_minutes,
                'idle_time_minutes': idle_time_minutes,
                'planned_runtime_minutes': planned_runtime
            })
        
        # Sort by date and limit
        all_records.sort(key=lambda x: x['record_date'] or '', reverse=True)
        all_records = all_records[:limit]
        
        return jsonify({
            'records': all_records
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/export-excel', methods=['GET'])
@jwt_required()
def export_controller_excel():
    """Export Controller report to Excel - per tanggal per mesin per work order"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from models.production import ShiftProduction
        from collections import Counter
        
        machine_id = request.args.get('machine_id', type=int)
        period = request.args.get('period', 'day')  # 'day', 'week' or 'month'
        selected_date = request.args.get('date')  # Optional: specific date for 'day' period
        
        # Calculate date range based on period
        if period == 'day':
            if selected_date:
                # Use the selected date
                start_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
                end_date = start_date
            else:
                # Find the latest production date for this machine
                latest_query = ShiftProduction.query
                if machine_id:
                    latest_query = latest_query.filter(ShiftProduction.machine_id == machine_id)
                latest_record = latest_query.order_by(ShiftProduction.production_date.desc()).first()
                
                if latest_record:
                    start_date = latest_record.production_date
                    end_date = latest_record.production_date
                else:
                    start_date = datetime.now().date()
                    end_date = datetime.now().date()
        else:
            end_date = datetime.now().date()
            if period == 'month':
                start_date = end_date - timedelta(days=30)
            else:  # week
                start_date = end_date - timedelta(days=7)
        
        # Get shift production records within date range
        query = ShiftProduction.query.filter(
            ShiftProduction.production_date >= start_date,
            ShiftProduction.production_date <= end_date
        )
        if machine_id:
            query = query.filter(ShiftProduction.machine_id == machine_id)
        
        records = query.order_by(
            ShiftProduction.production_date.desc(),
            ShiftProduction.machine_id,
            ShiftProduction.work_order_id
        ).all()
        
        # Keywords for human-related downtime (excluded from top 3)
        HUMAN_DOWNTIME_KEYWORDS = [
            'istirahat', 'makan', 'sholat', 'toilet', 'wc', 'break',
            'pulang', 'datang', 'terlambat', 'absen', 'cuti', 'sakit',
            'meeting', 'rapat', 'briefing', 'training', 'pelatihan'
        ]
        
        def is_human_downtime(reason):
            """Check if downtime reason is human-related"""
            reason_lower = reason.lower()
            return any(keyword in reason_lower for keyword in HUMAN_DOWNTIME_KEYWORDS)
        
        # Collect all issues across all records to find most frequent
        all_downtime_issues = []
        for sp in records:
            if sp.issues:
                parts = [p.strip() for p in sp.issues.split(';') if p.strip()]
                for part in parts:
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+?)(?:\s*\[.+\])?$', part, re.IGNORECASE)
                    if match:
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[.+\]\s*$', '', reason).strip()
                        duration = int(match.group(1))
                        all_downtime_issues.append({
                            'reason': reason,
                            'duration': duration,
                            'is_human': is_human_downtime(reason)
                        })
        
        # Count frequency of each issue (excluding human downtime for top ranking)
        non_human_issues = [i['reason'] for i in all_downtime_issues if not i['is_human']]
        issue_frequency = Counter(non_human_issues)
        
        # Get top 3 most frequent non-human issues
        top_3_issues = [issue for issue, count in issue_frequency.most_common(3)]
        print(f"[DEBUG] Top 3 non-human issues: {top_3_issues}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Controller Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Styles for downtime
        red_font = Font(color="FF0000", bold=True)  # Red for top 3 issues
        black_font = Font(color="000000")  # Black for other issues
        
        # Headers - single Issue column, issues go down as rows
        headers = [
            "Tanggal", "Mesin", "Shift", "Produk", 
            "Target (pcs)", "Aktual (pcs)", "Target (karton)", "Aktual (karton)",
            "Grade A", "Grade B", "Grade C",
            "Availability (%)", "Performance (%)", "Quality (%)", "OEE (%)",
            "Downtime (menit)", "Issue"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        row_num = 2
        for sp in records:
            # Get product name
            product_name = None
            if sp.product:
                product_name = sp.product.name
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
            
            # Get pack_per_karton
            pack_per_karton = 50
            if product_name:
                try:
                    search_name = product_name.replace('WIP ', '').strip()
                    product_new = ProductNew.query.filter(
                        ProductNew.nama_produk.ilike(f'%{search_name}%')
                    ).first()
                    if product_new and product_new.pack_per_karton:
                        pack_per_karton = int(product_new.pack_per_karton)
                except:
                    pass
            
            # Get shift number
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Parse all issues from issues field
            record_issues = []
            if sp.issues:
                parts = [p.strip() for p in sp.issues.split(';') if p.strip()]
                for part in parts:
                    # Parse format: "XX menit - reason [category]"
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+?)(?:\s*\[.+\])?$', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[.+\]\s*$', '', reason).strip()
                        is_top3 = reason in top_3_issues  # Check if in global top 3
                        record_issues.append({
                            'duration': duration, 
                            'reason': reason,
                            'is_top3': is_top3,
                            'is_human': is_human_downtime(reason)
                        })
                    else:
                        # Fallback: just use the text
                        record_issues.append({
                            'duration': 0, 
                            'reason': part,
                            'is_top3': False,
                            'is_human': is_human_downtime(part)
                        })
                
                # Sort: top 3 issues first (by duration), then others (by duration)
                record_issues.sort(key=lambda x: (not x['is_top3'], -x['duration']))
            
            target_qty = int(sp.target_quantity) if sp.target_quantity else 0
            actual_qty = int(sp.actual_quantity) if sp.actual_quantity else 0
            
            # Base row data (without issues)
            base_row_data = [
                sp.production_date.strftime('%Y-%m-%d') if sp.production_date else '',
                sp.machine.name if sp.machine else '',
                f"Shift {shift_num}",
                product_name or '',
                target_qty,
                actual_qty,
                round(target_qty / pack_per_karton) if pack_per_karton else 0,
                round(actual_qty / pack_per_karton) if pack_per_karton else 0,
                int(sp.good_quantity) if sp.good_quantity else 0,
                int(sp.rework_quantity) if sp.rework_quantity else 0,
                int(sp.reject_quantity) if sp.reject_quantity else 0,
                round(float(sp.efficiency_rate), 1) if sp.efficiency_rate else 0,
                100.0,  # Default performance
                round(float(sp.quality_rate), 1) if sp.quality_rate else 0,
                round(float(sp.oee_score), 1) if sp.oee_score else 0,
                int(sp.downtime_minutes) if sp.downtime_minutes else 0
            ]
            
            # If no issues, write single row with empty issue
            if not record_issues:
                for col, value in enumerate(base_row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    if col >= 5:
                        cell.alignment = Alignment(horizontal="right")
                # Empty issue cell
                cell = ws.cell(row=row_num, column=17, value="")
                cell.border = thin_border
                row_num += 1
            else:
                # Write multiple rows - one per issue
                # First row has all data, subsequent rows only have issue
                for i, issue in enumerate(record_issues):
                    if i == 0:
                        # First row: write all base data + first issue
                        for col, value in enumerate(base_row_data, 1):
                            cell = ws.cell(row=row_num, column=col, value=value)
                            cell.border = thin_border
                            if col >= 5:
                                cell.alignment = Alignment(horizontal="right")
                    else:
                        # Subsequent rows: empty cells for base data (or merge later)
                        for col in range(1, 17):
                            cell = ws.cell(row=row_num, column=col, value="")
                            cell.border = thin_border
                    
                    # Write issue in column 17
                    issue_text = f"{issue['duration']} menit - {issue['reason']}"
                    cell = ws.cell(row=row_num, column=17, value=issue_text)
                    cell.border = thin_border
                    
                    # Top 3 global issues (non-human) in red, others in black
                    # Human downtime never gets red even if frequent
                    if issue['is_top3'] and not issue['is_human']:
                        cell.font = red_font
                    else:
                        cell.font = black_font
                    
                    row_num += 1
        
        # Set fixed column widths for better readability
        column_widths = {
            1: 12,   # Tanggal
            2: 15,   # Mesin
            3: 8,    # Shift
            4: 25,   # Produk
            5: 12,   # Target (pcs)
            6: 12,   # Aktual (pcs)
            7: 10,   # Target (karton)
            8: 10,   # Aktual (karton)
            9: 10,   # Grade A
            10: 10,  # Grade B
            11: 10,  # Grade C
            12: 12,  # Availability
            13: 12,  # Performance
            14: 10,  # Quality
            15: 8,   # OEE
            16: 12,  # Downtime
            17: 40   # Issue
        }
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"controller_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/records', methods=['POST'])
@jwt_required()
def create_record():
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        record_number = generate_number('OEE', OEERecord, 'record_number')
        
        # Calculate OEE metrics
        availability = ((data['planned_production_time'] - data['downtime']) / data['planned_production_time']) * 100
        performance = ((data['total_pieces_produced'] * data['ideal_cycle_time']) / data['actual_production_time']) * 100 if data['actual_production_time'] > 0 else 0
        quality = (data['good_pieces'] / data['total_pieces_produced']) * 100 if data['total_pieces_produced'] > 0 else 0
        oee = (availability * performance * quality) / 10000
        
        record = OEERecord(
            record_number=record_number,
            machine_id=data['machine_id'],
            work_order_id=data.get('work_order_id'),
            record_date=datetime.fromisoformat(data['record_date']),
            shift=data.get('shift'),
            planned_production_time=data['planned_production_time'],
            downtime=data['downtime'],
            actual_production_time=data['actual_production_time'],
            ideal_cycle_time=data['ideal_cycle_time'],
            total_pieces_produced=data['total_pieces_produced'],
            good_pieces=data['good_pieces'],
            rejected_pieces=data['rejected_pieces'],
            availability=availability,
            performance=performance,
            quality=quality,
            oee=oee,
            recorded_by=user_id
        )
        
        db.session.add(record)
        db.session.commit()
        return jsonify({'message': 'OEE record created', 'record_id': record.id, 'oee': oee}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/downtime', methods=['GET'])
@jwt_required()
def get_downtime():
    """Get downtime records from both OEEDowntimeRecord and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        machine_id = request.args.get('machine_id', type=int)
        limit = request.args.get('limit', 50, type=int)
        
        all_records = []
        
        # Get from OEEDowntimeRecord
        query = OEEDowntimeRecord.query
        if machine_id:
            query = query.filter(OEEDowntimeRecord.machine_id == machine_id)
        
        oee_downtimes = query.order_by(OEEDowntimeRecord.start_time.desc()).limit(limit).all()
        for r in oee_downtimes:
            all_records.append({
                'id': r.id,
                'source': 'oee_downtime',
                'machine_id': r.machine_id,
                'reason': r.reason or r.downtime_category or 'Unknown',
                'duration_minutes': r.duration_minutes or 0,
                'start_time': r.start_time.isoformat() if r.start_time else None,
                'end_time': r.end_time.isoformat() if r.end_time else None,
                'downtime_category': r.downtime_category
            })
        
        # Get from ShiftProduction - create downtime entries per category
        shift_query = ShiftProduction.query
        if machine_id:
            shift_query = shift_query.filter(ShiftProduction.machine_id == machine_id)
        
        shift_records = shift_query.order_by(ShiftProduction.production_date.desc()).limit(limit).all()
        for sp in shift_records:
            base_time = datetime.combine(sp.production_date, sp.shift_start) if sp.production_date and sp.shift_start else datetime.now()
            
            # Add downtime entries for each category that has minutes
            if sp.downtime_mesin and sp.downtime_mesin > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-mesin",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Mesin',
                    'duration_minutes': sp.downtime_mesin,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'mesin'
                })
            
            if sp.downtime_operator and sp.downtime_operator > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-operator",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Operator',
                    'duration_minutes': sp.downtime_operator,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'operator'
                })
            
            if sp.downtime_material and sp.downtime_material > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-material",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Material',
                    'duration_minutes': sp.downtime_material,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'material'
                })
            
            if sp.downtime_design and sp.downtime_design > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-design",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Design Change',
                    'duration_minutes': sp.downtime_design,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'design'
                })
            
            if sp.downtime_others and sp.downtime_others > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-others",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': sp.issues or 'Downtime Lainnya',
                    'duration_minutes': sp.downtime_others,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'others'
                })
        
        # Sort by start_time and limit
        all_records.sort(key=lambda x: x['start_time'] or '', reverse=True)
        all_records = all_records[:limit]
        
        return jsonify({
            'records': all_records
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/downtime', methods=['POST'])
@jwt_required()
def create_downtime():
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        downtime = OEEDowntimeRecord(
            oee_record_id=data.get('oee_record_id'),
            machine_id=data['machine_id'],
            start_time=datetime.fromisoformat(data['start_time']),
            end_time=datetime.fromisoformat(data['end_time']) if data.get('end_time') else None,
            duration_minutes=data.get('duration_minutes'),
            downtime_category=data['downtime_category'],
            reason=data.get('reason'),
            recorded_by=user_id
        )
        
        db.session.add(downtime)
        db.session.commit()
        return jsonify(success_response('api.success')), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/shift-production/<int:id>/downtime', methods=['PUT'])
@jwt_required()
def update_shift_production_downtime(id):
    """Update downtime breakdown for a ShiftProduction record"""
    try:
        from models.production import ShiftProduction
        
        sp = ShiftProduction.query.get(id)
        if not sp:
            return jsonify({'error': 'ShiftProduction not found'}), 404
        
        data = request.get_json()
        
        # Update downtime by category
        if 'downtime_mesin' in data:
            sp.downtime_mesin = int(data['downtime_mesin'])
        if 'downtime_operator' in data:
            sp.downtime_operator = int(data['downtime_operator'])
        if 'downtime_material' in data:
            sp.downtime_material = int(data['downtime_material'])
        if 'downtime_design' in data:
            sp.downtime_design = int(data['downtime_design'])
        if 'downtime_others' in data:
            sp.downtime_others = int(data['downtime_others'])
        
        # Recalculate total downtime
        sp.downtime_minutes = (sp.downtime_mesin or 0) + (sp.downtime_operator or 0) + \
                              (sp.downtime_material or 0) + (sp.downtime_design or 0) + \
                              (sp.downtime_others or 0)
        
        # Recalculate loss percentages
        planned_runtime = sp.planned_runtime or 480
        sp.loss_mesin = round((sp.downtime_mesin or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_operator = round((sp.downtime_operator or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_material = round((sp.downtime_material or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_design = round((sp.downtime_design or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_others = round((sp.downtime_others or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        
        # Recalculate efficiency and OEE
        sp.actual_runtime = planned_runtime - sp.downtime_minutes
        sp.efficiency_rate = round((sp.actual_runtime / planned_runtime * 100) if planned_runtime > 0 else 100, 2)
        sp.oee_score = round((sp.efficiency_rate * float(sp.quality_rate or 100)) / 100, 2)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Downtime updated successfully',
            'data': {
                'id': sp.id,
                'downtime_mesin': sp.downtime_mesin,
                'downtime_operator': sp.downtime_operator,
                'downtime_material': sp.downtime_material,
                'downtime_design': sp.downtime_design,
                'downtime_others': sp.downtime_others,
                'downtime_minutes': sp.downtime_minutes,
                'efficiency_rate': float(sp.efficiency_rate),
                'oee_score': float(sp.oee_score)
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/shift-production', methods=['GET'])
@jwt_required()
def get_shift_productions():
    """Get ShiftProduction records for editing"""
    try:
        from models.production import ShiftProduction
        
        machine_id = request.args.get('machine_id', type=int)
        limit = request.args.get('limit', 50, type=int)
        
        query = ShiftProduction.query
        if machine_id:
            query = query.filter(ShiftProduction.machine_id == machine_id)
        
        records = query.order_by(ShiftProduction.production_date.desc()).limit(limit).all()
        
        return jsonify({
            'records': [{
                'id': sp.id,
                'production_date': sp.production_date.isoformat() if sp.production_date else None,
                'shift': sp.shift,
                'machine_id': sp.machine_id,
                'machine_name': sp.machine.name if sp.machine else None,
                'product_name': sp.product.name if sp.product else None,
                'actual_quantity': float(sp.actual_quantity) if sp.actual_quantity else 0,
                'downtime_minutes': sp.downtime_minutes or 0,
                'downtime_mesin': sp.downtime_mesin or 0,
                'downtime_operator': sp.downtime_operator or 0,
                'downtime_material': sp.downtime_material or 0,
                'downtime_design': sp.downtime_design or 0,
                'downtime_others': sp.downtime_others or 0,
                'efficiency_rate': float(sp.efficiency_rate) if sp.efficiency_rate else 0,
                'quality_rate': float(sp.quality_rate) if sp.quality_rate else 0,
                'oee_score': float(sp.oee_score) if sp.oee_score else 0,
                'issues': sp.issues
            } for sp in records]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===============================
# ENHANCED OEE ENDPOINTS
# ===============================

@oee_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_oee_dashboard():
    """Get comprehensive OEE dashboard data from both OEERecord and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        # Get query parameters
        machine_id_param = request.args.get('machine_id')
        machine_id = None
        if machine_id_param and machine_id_param != 'null' and machine_id_param != '':
            try:
                machine_id = int(machine_id_param)
            except (ValueError, TypeError):
                machine_id = None
        
        days = request.args.get('days', 30, type=int)
        
        # Date range
        end_date = date.today()
        start_date = end_date - timedelta(days=days)
        
        # Get OEERecord data
        oee_query = OEERecord.query.filter(
            OEERecord.record_date.between(start_date, end_date)
        )
        if machine_id:
            oee_query = oee_query.filter(OEERecord.machine_id == machine_id)
        oee_records = oee_query.all()
        
        # Get ShiftProduction data (from WorkOrder production input)
        shift_query = ShiftProduction.query.filter(
            ShiftProduction.production_date.between(start_date, end_date)
        )
        if machine_id:
            shift_query = shift_query.filter(ShiftProduction.machine_id == machine_id)
        shift_records = shift_query.all()
        
        # Combine records for calculations
        records = oee_records  # Keep original for compatibility
        
        # Calculate overall metrics combining both sources
        all_oee_values = []
        all_availability_values = []
        all_performance_values = []
        all_quality_values = []
        
        # From OEERecord
        for r in oee_records:
            all_oee_values.append(float(r.oee) if r.oee else 0)
            all_availability_values.append(float(r.availability) if r.availability else 0)
            all_performance_values.append(float(r.performance) if r.performance else 0)
            all_quality_values.append(float(r.quality) if r.quality else 0)
        
        # From ShiftProduction
        for sp in shift_records:
            all_oee_values.append(float(sp.oee_score) if sp.oee_score else 0)
            all_availability_values.append(float(sp.efficiency_rate) if sp.efficiency_rate else 0)  # efficiency as availability proxy
            all_performance_values.append(100.0)  # Default performance
            all_quality_values.append(float(sp.quality_rate) if sp.quality_rate else 0)
        
        if all_oee_values:
            avg_oee = sum(all_oee_values) / len(all_oee_values)
            avg_availability = sum(all_availability_values) / len(all_availability_values)
            avg_performance = sum(all_performance_values) / len(all_performance_values)
            avg_quality = sum(all_quality_values) / len(all_quality_values)
            best_oee = max(all_oee_values)
            worst_oee = min(all_oee_values)
        else:
            avg_oee = avg_availability = avg_performance = avg_quality = 0
            best_oee = worst_oee = 0
        
        # Get machine performance data
        machines_query = Machine.query.filter(Machine.is_active == True)
        if machine_id:
            machines_query = machines_query.filter(Machine.id == machine_id)
        
        machines = machines_query.all()
        machine_performance = []
        
        for machine in machines:
            # From OEERecord
            machine_oee_records = [r for r in oee_records if r.machine_id == machine.id]
            # From ShiftProduction
            machine_shift_records = [sp for sp in shift_records if sp.machine_id == machine.id]
            
            # Combine OEE values
            machine_oee_values = []
            machine_downtime = 0
            machine_production = 0
            
            for r in machine_oee_records:
                machine_oee_values.append(float(r.oee) if r.oee else 0)
                machine_downtime += r.downtime or 0
                machine_production += r.total_pieces_produced or 0
            
            for sp in machine_shift_records:
                machine_oee_values.append(float(sp.oee_score) if sp.oee_score else 0)
                machine_downtime += sp.downtime_minutes or 0
                machine_production += float(sp.actual_quantity) if sp.actual_quantity else 0
            
            if machine_oee_values:
                machine_avg_oee = sum(machine_oee_values) / len(machine_oee_values)
            else:
                machine_avg_oee = 0
            
            # Get maintenance info
            next_maintenance = MaintenanceSchedule.query.filter(
                MaintenanceSchedule.machine_id == machine.id,
                MaintenanceSchedule.is_active == True,
                MaintenanceSchedule.next_maintenance_date >= date.today()
            ).order_by(MaintenanceSchedule.next_maintenance_date).first()
            
            # Get recent alerts
            recent_alerts = OEEAlert.query.filter(
                OEEAlert.machine_id == machine.id,
                OEEAlert.status == 'active'
            ).count()
            
            machine_performance.append({
                'machine_id': machine.id,
                'machine_name': machine.name,
                'machine_code': machine.code,
                'status': machine.status,
                'avg_oee': round(machine_avg_oee, 2),
                'total_downtime': machine_downtime,
                'total_production': machine_production,
                'next_maintenance': next_maintenance.next_maintenance_date.isoformat() if next_maintenance else None,
                'active_alerts': recent_alerts,
                'efficiency': float(machine.efficiency) if machine.efficiency else 100,
                'availability': float(machine.availability) if machine.availability else 100
            })
        
        # Get trend data (last 7 days) - combining both sources
        trend_data = []
        for i in range(7):
            trend_date = end_date - timedelta(days=i)
            day_oee_values = []
            
            # From OEERecord
            for r in oee_records:
                if r.record_date == trend_date:
                    day_oee_values.append(float(r.oee) if r.oee else 0)
            
            # From ShiftProduction
            for sp in shift_records:
                if sp.production_date == trend_date:
                    day_oee_values.append(float(sp.oee_score) if sp.oee_score else 0)
            
            if day_oee_values:
                day_avg_oee = sum(day_oee_values) / len(day_oee_values)
            else:
                day_avg_oee = 0
            
            trend_data.append({
                'date': trend_date.isoformat(),
                'oee': round(day_avg_oee, 2)
            })
        
        trend_data.reverse()  # Show oldest to newest
        
        # Get active alerts
        alerts_query = OEEAlert.query.filter(OEEAlert.status == 'active')
        if machine_id:
            alerts_query = alerts_query.filter(OEEAlert.machine_id == machine_id)
        
        active_alerts = alerts_query.order_by(desc(OEEAlert.alert_date)).limit(10).all()
        
        # Get downtime analysis - combining both sources
        downtime_by_category = {}
        
        # From OEEDowntimeRecord
        downtime_records = OEEDowntimeRecord.query.join(OEERecord).filter(
            OEERecord.record_date.between(start_date, end_date)
        )
        if machine_id:
            downtime_records = downtime_records.filter(OEEDowntimeRecord.machine_id == machine_id)
        
        for downtime in downtime_records.all():
            category = downtime.downtime_category or 'others'
            if category not in downtime_by_category:
                downtime_by_category[category] = 0
            downtime_by_category[category] += downtime.duration_minutes or 0
        
        # From ShiftProduction - add downtime by category
        for sp in shift_records:
            if sp.downtime_mesin and sp.downtime_mesin > 0:
                downtime_by_category['mesin'] = downtime_by_category.get('mesin', 0) + sp.downtime_mesin
            if sp.downtime_operator and sp.downtime_operator > 0:
                downtime_by_category['operator'] = downtime_by_category.get('operator', 0) + sp.downtime_operator
            if sp.downtime_material and sp.downtime_material > 0:
                downtime_by_category['material'] = downtime_by_category.get('material', 0) + sp.downtime_material
            if sp.downtime_design and sp.downtime_design > 0:
                downtime_by_category['design'] = downtime_by_category.get('design', 0) + sp.downtime_design
            if sp.downtime_others and sp.downtime_others > 0:
                downtime_by_category['others'] = downtime_by_category.get('others', 0) + sp.downtime_others
        
        return jsonify({
            'summary': {
                'avg_oee': round(avg_oee, 2),
                'avg_availability': round(avg_availability, 2),
                'avg_performance': round(avg_performance, 2),
                'avg_quality': round(avg_quality, 2),
                'best_oee': round(best_oee, 2),
                'worst_oee': round(worst_oee, 2),
                'total_records': len(oee_records) + len(shift_records),
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            },
            'machine_performance': machine_performance,
            'trend_data': trend_data,
            'active_alerts': [{
                'id': alert.id,
                'machine_name': alert.machine.name,
                'alert_type': alert.alert_type,
                'severity': alert.severity,
                'title': alert.title,
                'message': alert.message,
                'alert_date': alert.alert_date.isoformat(),
                'threshold_value': float(alert.threshold_value) if alert.threshold_value else None,
                'actual_value': float(alert.actual_value) if alert.actual_value else None
            } for alert in active_alerts],
            'downtime_analysis': [
                {'category': category, 'minutes': minutes}
                for category, minutes in downtime_by_category.items()
            ]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/alerts', methods=['GET'])
@jwt_required()
def get_alerts():
    """Get OEE alerts"""
    try:
        status = request.args.get('status', 'active')
        
        # Handle machine_id parameter properly
        machine_id_param = request.args.get('machine_id')
        machine_id = None
        if machine_id_param and machine_id_param != 'null' and machine_id_param != '':
            try:
                machine_id = int(machine_id_param)
            except (ValueError, TypeError):
                machine_id = None
        
        severity = request.args.get('severity')
        
        query = OEEAlert.query
        
        if status:
            query = query.filter(OEEAlert.status == status)
        if machine_id:
            query = query.filter(OEEAlert.machine_id == machine_id)
        if severity:
            query = query.filter(OEEAlert.severity == severity)
        
        alerts = query.order_by(desc(OEEAlert.alert_date)).all()
        
        return jsonify({
            'alerts': [{
                'id': alert.id,
                'machine_id': alert.machine_id,
                'machine_name': alert.machine.name,
                'alert_type': alert.alert_type,
                'severity': alert.severity,
                'title': alert.title,
                'message': alert.message,
                'threshold_value': float(alert.threshold_value) if alert.threshold_value else None,
                'actual_value': float(alert.actual_value) if alert.actual_value else None,
                'alert_date': alert.alert_date.isoformat(),
                'status': alert.status,
                'acknowledged_by': alert.acknowledged_by_user.username if alert.acknowledged_by_user else None,
                'acknowledged_at': alert.acknowledged_at.isoformat() if alert.acknowledged_at else None,
                'resolved_by': alert.resolved_by_user.username if alert.resolved_by_user else None,
                'resolved_at': alert.resolved_at.isoformat() if alert.resolved_at else None
            } for alert in alerts]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/alerts/<int:alert_id>/acknowledge', methods=['PUT'])
@jwt_required()
def acknowledge_alert(alert_id):
    """Acknowledge an OEE alert"""
    try:
        user_id = get_jwt_identity()
        alert = OEEAlert.query.get_or_404(alert_id)
        
        alert.status = 'acknowledged'
        alert.acknowledged_by = user_id
        alert.acknowledged_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify(success_response('api.success')), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/alerts/<int:alert_id>/resolve', methods=['PUT'])
@jwt_required()
def resolve_alert(alert_id):
    """Resolve an OEE alert"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        alert = OEEAlert.query.get_or_404(alert_id)
        
        alert.status = 'resolved'
        alert.resolved_by = user_id
        alert.resolved_at = datetime.utcnow()
        alert.resolution_notes = data.get('resolution_notes', '')
        
        db.session.commit()
        
        return jsonify(success_response('api.success')), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/maintenance-impact', methods=['POST'])
@jwt_required()
def create_maintenance_impact():
    """Create maintenance impact record"""
    try:
        data = request.get_json()
        
        impact = MaintenanceImpact(
            maintenance_record_id=data['maintenance_record_id'],
            machine_id=data['machine_id'],
            impact_date=datetime.strptime(data['impact_date'], '%Y-%m-%d').date(),
            planned_downtime_hours=data.get('planned_downtime_hours', 0),
            actual_downtime_hours=data.get('actual_downtime_hours', 0),
            production_loss_units=data.get('production_loss_units', 0),
            revenue_impact=data.get('revenue_impact', 0),
            oee_before_maintenance=data.get('oee_before_maintenance'),
            oee_after_maintenance=data.get('oee_after_maintenance'),
            notes=data.get('notes')
        )
        
        # Calculate improvement percentage
        if impact.oee_before_maintenance and impact.oee_after_maintenance:
            impact.improvement_percentage = float(impact.oee_after_maintenance) - float(impact.oee_before_maintenance)
        
        db.session.add(impact)
        db.session.commit()
        
        return jsonify({
            'message': 'Maintenance impact recorded successfully',
            'impact_id': impact.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/machines/<int:machine_id>/analytics', methods=['GET'])
@jwt_required()
def get_machine_analytics(machine_id):
    """Get detailed analytics for a specific machine"""
    try:
        machine = Machine.query.get_or_404(machine_id)
        
        # Get query parameters
        period = request.args.get('period', 'monthly')  # daily, weekly, monthly
        months = request.args.get('months', 6, type=int)
        
        # Get analytics data
        analytics = OEEAnalytics.query.filter(
            OEEAnalytics.machine_id == machine_id,
            OEEAnalytics.period_type == period
        ).order_by(desc(OEEAnalytics.analysis_date)).limit(months).all()
        
        # Get maintenance impact data
        maintenance_impacts = MaintenanceImpact.query.filter(
            MaintenanceImpact.machine_id == machine_id
        ).order_by(desc(MaintenanceImpact.impact_date)).limit(10).all()
        
        # Get recent OEE records for detailed view
        recent_records = OEERecord.query.filter(
            OEERecord.machine_id == machine_id
        ).order_by(desc(OEERecord.record_date)).limit(30).all()
        
        # Calculate trends
        if len(analytics) >= 2:
            latest = analytics[0]
            previous = analytics[1]
            oee_trend = float(latest.avg_oee) - float(previous.avg_oee)
            availability_trend = float(latest.avg_availability) - float(previous.avg_availability)
            performance_trend = float(latest.avg_performance) - float(previous.avg_performance)
            quality_trend = float(latest.avg_quality) - float(previous.avg_quality)
        else:
            oee_trend = availability_trend = performance_trend = quality_trend = 0
        
        return jsonify({
            'machine': {
                'id': machine.id,
                'name': machine.name,
                'code': machine.code,
                'type': machine.machine_type,
                'status': machine.status,
                'capacity_per_hour': float(machine.capacity_per_hour) if machine.capacity_per_hour else None,
                'last_maintenance': machine.last_maintenance.isoformat() if machine.last_maintenance else None,
                'next_maintenance': machine.next_maintenance.isoformat() if machine.next_maintenance else None
            },
            'trends': {
                'oee_trend': round(oee_trend, 2),
                'availability_trend': round(availability_trend, 2),
                'performance_trend': round(performance_trend, 2),
                'quality_trend': round(quality_trend, 2)
            },
            'analytics': [{
                'date': a.analysis_date.isoformat(),
                'period_type': a.period_type,
                'avg_oee': float(a.avg_oee),
                'avg_availability': float(a.avg_availability),
                'avg_performance': float(a.avg_performance),
                'avg_quality': float(a.avg_quality),
                'total_downtime_hours': float(a.total_downtime_hours),
                'total_production_hours': float(a.total_production_hours),
                'total_units_produced': float(a.total_units_produced),
                'defect_rate': float(a.defect_rate),
                'maintenance_hours': float(a.maintenance_hours),
                'breakdown_count': a.breakdown_count
            } for a in analytics],
            'maintenance_impacts': [{
                'date': mi.impact_date.isoformat(),
                'planned_downtime': float(mi.planned_downtime_hours),
                'actual_downtime': float(mi.actual_downtime_hours),
                'production_loss': float(mi.production_loss_units),
                'revenue_impact': float(mi.revenue_impact),
                'oee_before': float(mi.oee_before_maintenance) if mi.oee_before_maintenance else None,
                'oee_after': float(mi.oee_after_maintenance) if mi.oee_after_maintenance else None,
                'improvement': float(mi.improvement_percentage) if mi.improvement_percentage else None
            } for mi in maintenance_impacts],
            'recent_records': [{
                'date': r.record_date.isoformat(),
                'shift': r.shift,
                'oee': float(r.oee),
                'availability': float(r.availability),
                'performance': float(r.performance),
                'quality': float(r.quality),
                'downtime': r.downtime,
                'total_pieces': r.total_pieces_produced,
                'good_pieces': r.good_pieces
            } for r in recent_records]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/daily-controller', methods=['GET'])
@jwt_required()
def get_daily_controller():
    """Get all machines' production data for a specific date - Daily Controller view"""
    try:
        from models.production import ShiftProduction
        
        selected_date = request.args.get('date')
        if selected_date:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            target_date = date.today()
        
        # Get all shift productions for this date
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date == target_date
        ).order_by(ShiftProduction.machine_id, ShiftProduction.shift).all()
        
        # Group by machine
        machines_data = {}
        
        # IDLE TIME keywords
        idle_keywords = [
            'tunggu kain', 'tunggu stiker', 'tunggu packaging', 'tunggu mixing',
            'tunggu bahan', 'tunggu material', 'tunggu label', 'tunggu box',
            'tunggu karton', 'tunggu lem', 'tunggu tinta', 'tunggu order',
            'menunggu kain', 'menunggu stiker', 'menunggu packaging', 'menunggu mixing',
            'nunggu kain', 'nunggu stiker', 'nunggu packaging', 'nunggu mixing',
            'waiting for', 'standby'
        ]
        
        for sp in shift_records:
            machine_id = sp.machine_id
            
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'date': target_date.isoformat(),
                    'shifts': [],
                    'total_planned': 0,
                    'total_runtime': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_output': 0,
                    'total_grade_a': 0,
                    'total_grade_b': 0,
                    'total_grade_c': 0,
                    'total_machine_speed': 0,
                    'total_target': 0,
                    'products': [],
                    'top_3_downtime': [],
                    'efficiency': 0,
                    'machine_efficiency': 0,
                    'quality': 0
                }
            
            # Parse downtime from issues - IDLE is SEPARATE from downtime
            downtime_breakdown = []
            shift_downtime = 0  # Non-idle downtime only
            shift_idle = 0  # Idle time separate
            
            if sp.issues:
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason).strip()
                        reason_lower = reason.lower()
                        
                        is_idle = any(kw in reason_lower for kw in idle_keywords)
                        if is_idle:
                            shift_idle += duration
                        else:
                            shift_downtime += duration
                        
                        downtime_breakdown.append({'reason': reason, 'duration_minutes': duration})
            
            # Fallback for downtime (only if parsing yielded 0)
            if shift_downtime == 0 and shift_idle == 0:
                if sp.downtime_minutes:
                    shift_downtime = int(sp.downtime_minutes)
                else:
                    shift_downtime = (
                        (int(sp.downtime_mesin) if sp.downtime_mesin else 0) +
                        (int(sp.downtime_operator) if sp.downtime_operator else 0) +
                        (int(sp.downtime_material) if sp.downtime_material else 0) +
                        (int(sp.downtime_design) if sp.downtime_design else 0) +
                        (int(sp.downtime_others) if sp.downtime_others else 0)
                    )
            
            # Extract shift number
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Get planned runtime per shift
            if shift_num == 1:
                default_planned = 510  # 06:30-15:00 = 8.5 hours
            elif shift_num == 2:
                default_planned = 480  # 15:00-23:00 = 8 hours
            else:
                default_planned = 450  # 23:00-06:30 = 7.5 hours
            
            planned_runtime = int(sp.planned_runtime) if sp.planned_runtime else default_planned
            
            # Runtime = Planned - Downtime - Idle (calculated)
            shift_runtime = max(0, planned_runtime - shift_downtime - shift_idle)
            
            # Get product name and packs_per_karton
            product_name = None
            pack_per_carton = 0
            if sp.product:
                product_name = sp.product.name
                if sp.product.packaging:
                    pack_per_carton = sp.product.packaging.packs_per_karton or 0
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
                if sp.work_order.product.packaging:
                    pack_per_carton = sp.work_order.product.packaging.packs_per_karton or 0
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            import math
            shift_data = {
                'shift': shift_num,
                'runtime_minutes': shift_runtime,
                'downtime_minutes': shift_downtime,
                'idle_time_minutes': shift_idle,
                'grade_a': grade_a,
                'grade_a_carton': math.floor(grade_a / pack_per_carton) if pack_per_carton > 0 else 0,
                'grade_b': int(sp.rework_quantity) if sp.rework_quantity else 0,
                'grade_c': int(sp.reject_quantity) if sp.reject_quantity else 0,
                'total': int(sp.actual_quantity) if sp.actual_quantity else 0,
                'product_name': product_name,
                'pack_per_carton': pack_per_carton,
                'wo_number': sp.work_order.wo_number if sp.work_order else None
            }
            
            # Get machine_speed from ShiftProduction
            shift_machine_speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            # Get target quantity from ShiftProduction or WorkOrder
            shift_target = int(sp.target_quantity) if sp.target_quantity else 0
            if shift_target == 0 and sp.work_order:
                shift_target = int(sp.work_order.quantity) if sp.work_order.quantity else 0
            
            machines_data[machine_id]['shifts'].append(shift_data)
            machines_data[machine_id]['total_planned'] += planned_runtime
            machines_data[machine_id]['total_runtime'] += shift_runtime
            machines_data[machine_id]['total_downtime'] += shift_downtime
            machines_data[machine_id]['total_idle'] += shift_idle
            machines_data[machine_id]['total_output'] += shift_data['total']
            machines_data[machine_id]['total_grade_a'] += shift_data['grade_a']
            machines_data[machine_id]['total_grade_b'] += shift_data['grade_b']
            machines_data[machine_id]['total_grade_c'] += shift_data['grade_c']
            machines_data[machine_id]['total_machine_speed'] += shift_machine_speed
            machines_data[machine_id]['total_target'] += shift_target
            
            if product_name and product_name not in machines_data[machine_id]['products']:
                machines_data[machine_id]['products'].append(product_name)
            
            # Accumulate downtime breakdown
            for dt in downtime_breakdown:
                existing = next((d for d in machines_data[machine_id]['top_3_downtime'] if d['reason'] == dt['reason']), None)
                if existing:
                    existing['duration_minutes'] += dt['duration_minutes']
                else:
                    machines_data[machine_id]['top_3_downtime'].append(dt.copy())
        
        # Keywords to exclude from Top 3 Downtime (biological/personal breaks)
        excluded_keywords = [
            'istirahat', 'sholat', 'solat', 'makan', 'minum', 'toilet', 
            'wc', 'buang air', 'pipis', 'bab', 'break', 'rest', 'pray',
            'ibadah', 'jumatan', 'jumat', 'dhuhur', 'ashar', 'maghrib'
        ]
        
        # Sort top 3 downtime and calculate efficiency for each machine
        for machine_id in machines_data:
            # Filter out biological/personal breaks from top downtime
            filtered_downtime = [
                dt for dt in machines_data[machine_id]['top_3_downtime']
                if not any(kw in dt['reason'].lower() for kw in excluded_keywords)
            ]
            filtered_downtime.sort(key=lambda x: x['duration_minutes'], reverse=True)
            machines_data[machine_id]['top_3_downtime'] = filtered_downtime[:3]
            # Sort shifts by shift number
            machines_data[machine_id]['shifts'].sort(key=lambda x: x['shift'])
            
            # Calculate MRT (Machine Run Time) = Grade A / machine speed
            total_grade_a = machines_data[machine_id]['total_grade_a']
            total_downtime = machines_data[machine_id]['total_downtime']
            total_idle = machines_data[machine_id]['total_idle']
            
            # Get machine speed from ShiftProduction (sum of all shifts, take average if multiple)
            # Speed is already in pcs/menit, no conversion needed
            machine_speed_total = machines_data[machine_id].get('total_machine_speed', 0)
            shift_count = len(machines_data[machine_id]['shifts'])
            machine_speed_per_minute = machine_speed_total / shift_count if shift_count > 0 else 0
            
            mrt = 0
            if machine_speed_per_minute > 0:
                mrt = total_grade_a / machine_speed_per_minute  # MRT in minutes
            
            # Total Time = MRT + Downtime + Idle
            total_time = mrt + total_downtime + total_idle
            
            # Efficiency = MRT / Total Time * 100
            if total_time > 0:
                machines_data[machine_id]['efficiency'] = round((mrt / total_time) * 100, 1)
            
            # Store MRT for display (integers only)
            mrt_int = int(round(mrt))
            machines_data[machine_id]['mrt'] = mrt_int
            # Total Waktu = Runtime + Downtime (not MRT-based)
            total_runtime = machines_data[machine_id]['total_runtime']
            machines_data[machine_id]['total_time'] = total_runtime + total_downtime + total_idle
            machines_data[machine_id]['machine_speed'] = int(round(machine_speed_per_minute))  # pcs/menit
            
            # Calculate Quality = (Grade A / Total Output) * 100
            total_output = machines_data[machine_id]['total_output']
            if total_output > 0:
                machines_data[machine_id]['quality'] = round((total_grade_a / total_output) * 100, 1)
            
            # Calculate Machine Efficiency = (Grade A / Target) * 100
            total_target = machines_data[machine_id]['total_target']
            if total_target > 0:
                machines_data[machine_id]['machine_efficiency'] = round((total_grade_a / total_target) * 100, 1)
        
        # Convert to list and sort by machine name
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        return jsonify({
            'date': target_date.isoformat(),
            'machines': result,
            'total_machines': len(result)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/efficiency-alerts', methods=['GET'])
@jwt_required()
def get_efficiency_alerts():
    """Get machines with efficiency below target for today"""
    try:
        from models.production import ShiftProduction, Machine
        
        # Get today's date
        today = datetime.now().date()
        
        # Get all shift productions for today
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date == today
        ).all()
        
        # Group by machine and calculate efficiency
        machines_data = {}
        
        for sp in shift_records:
            machine_id = sp.machine_id
            if not machine_id:
                continue
                
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'total_grade_a': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_machine_speed': 0,
                    'shift_count': 0
                }
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            downtime = int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            idle = int(sp.idle_time) if sp.idle_time else 0
            speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            machines_data[machine_id]['total_grade_a'] += grade_a
            machines_data[machine_id]['total_downtime'] += downtime
            machines_data[machine_id]['total_idle'] += idle
            machines_data[machine_id]['total_machine_speed'] += speed
            machines_data[machine_id]['shift_count'] += 1
        
        # Calculate efficiency and filter alerts
        alerts = []
        for machine_id, data in machines_data.items():
            speed_per_min = data['total_machine_speed'] / data['shift_count'] if data['shift_count'] > 0 else 0
            mrt = data['total_grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + data['total_downtime'] + data['total_idle']
            efficiency = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            
            target = data['target_efficiency']
            if efficiency < target:
                alert_level = 'critical' if efficiency < target * 0.7 else 'warning'
                alerts.append({
                    'machine_id': machine_id,
                    'machine_name': data['machine_name'],
                    'efficiency': efficiency,
                    'target': target,
                    'gap': round(target - efficiency, 1),
                    'level': alert_level
                })
        
        # Sort by gap (largest gap first)
        alerts.sort(key=lambda x: x['gap'], reverse=True)
        
        return jsonify({
            'date': today.isoformat(),
            'alerts': alerts,
            'total_alerts': len(alerts),
            'critical_count': len([a for a in alerts if a['level'] == 'critical']),
            'warning_count': len([a for a in alerts if a['level'] == 'warning'])
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/weekly-controller', methods=['GET'])
@jwt_required()
def get_weekly_controller():
    """Get weekly efficiency summary for all machines"""
    try:
        from models.production import ShiftProduction, Machine
        
        # Get week start date (Monday) - default to current week
        date_str = request.args.get('week_start')
        if date_str:
            week_start = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            today = datetime.now().date()
            week_start = today - timedelta(days=today.weekday())  # Monday
        
        week_end = week_start + timedelta(days=6)  # Sunday
        
        # Get all shift productions for the week
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date >= week_start,
            ShiftProduction.production_date <= week_end
        ).all()
        
        # Group by machine
        machines_data = {}
        
        for sp in shift_records:
            machine_id = sp.machine_id
            if not machine_id:
                continue
                
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'daily_data': {},
                    'total_grade_a': 0,
                    'total_output': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_machine_speed': 0,
                    'shift_count': 0
                }
            
            # Group by date
            date_key = sp.production_date.isoformat()
            if date_key not in machines_data[machine_id]['daily_data']:
                machines_data[machine_id]['daily_data'][date_key] = {
                    'grade_a': 0,
                    'output': 0,
                    'downtime': 0,
                    'idle': 0,
                    'machine_speed': 0,
                    'shift_count': 0
                }
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            output = int(sp.actual_quantity) if sp.actual_quantity else 0
            downtime = int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            idle = int(sp.idle_time) if sp.idle_time else 0
            speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            machines_data[machine_id]['daily_data'][date_key]['grade_a'] += grade_a
            machines_data[machine_id]['daily_data'][date_key]['output'] += output
            machines_data[machine_id]['daily_data'][date_key]['downtime'] += downtime
            machines_data[machine_id]['daily_data'][date_key]['idle'] += idle
            machines_data[machine_id]['daily_data'][date_key]['machine_speed'] += speed
            machines_data[machine_id]['daily_data'][date_key]['shift_count'] += 1
            
            machines_data[machine_id]['total_grade_a'] += grade_a
            machines_data[machine_id]['total_output'] += output
            machines_data[machine_id]['total_downtime'] += downtime
            machines_data[machine_id]['total_idle'] += idle
            machines_data[machine_id]['total_machine_speed'] += speed
            machines_data[machine_id]['shift_count'] += 1
        
        # Calculate weekly efficiency for each machine
        for machine_id in machines_data:
            data = machines_data[machine_id]
            
            # Calculate daily efficiencies
            for date_key in data['daily_data']:
                day_data = data['daily_data'][date_key]
                speed_per_min = day_data['machine_speed'] / day_data['shift_count'] if day_data['shift_count'] > 0 else 0
                mrt = day_data['grade_a'] / speed_per_min if speed_per_min > 0 else 0
                total_time = mrt + day_data['downtime'] + day_data['idle']
                day_data['efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
                day_data['mrt'] = round(mrt, 1)
            
            # Calculate weekly average efficiency
            speed_per_min = data['total_machine_speed'] / data['shift_count'] if data['shift_count'] > 0 else 0
            mrt = data['total_grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + data['total_downtime'] + data['total_idle']
            data['avg_efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            data['mrt'] = round(mrt, 1)
            data['total_time'] = round(total_time, 1)
            data['quality'] = round((data['total_grade_a'] / data['total_output']) * 100, 1) if data['total_output'] > 0 else 0
        
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        return jsonify({
            'week_start': week_start.isoformat(),
            'week_end': week_end.isoformat(),
            'machines': result,
            'total_machines': len(result)
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/monthly-controller', methods=['GET'])
@jwt_required()
def get_monthly_controller():
    """Get monthly efficiency summary for all machines"""
    try:
        from models.production import ShiftProduction, Machine
        from calendar import monthrange
        
        # Get month - default to current month
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        
        month_start = date(year, month, 1)
        _, last_day = monthrange(year, month)
        month_end = date(year, month, last_day)
        
        # Get all shift productions for the month
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date >= month_start,
            ShiftProduction.production_date <= month_end
        ).all()
        
        # Group by machine
        machines_data = {}
        
        for sp in shift_records:
            machine_id = sp.machine_id
            if not machine_id:
                continue
                
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'weekly_data': {},
                    'total_grade_a': 0,
                    'total_output': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_machine_speed': 0,
                    'shift_count': 0
                }
            
            # Group by week number
            week_num = sp.production_date.isocalendar()[1]
            week_key = f'W{week_num}'
            if week_key not in machines_data[machine_id]['weekly_data']:
                machines_data[machine_id]['weekly_data'][week_key] = {
                    'grade_a': 0,
                    'output': 0,
                    'downtime': 0,
                    'idle': 0,
                    'machine_speed': 0,
                    'shift_count': 0
                }
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            output = int(sp.actual_quantity) if sp.actual_quantity else 0
            downtime = int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            idle = int(sp.idle_time) if sp.idle_time else 0
            speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            machines_data[machine_id]['weekly_data'][week_key]['grade_a'] += grade_a
            machines_data[machine_id]['weekly_data'][week_key]['output'] += output
            machines_data[machine_id]['weekly_data'][week_key]['downtime'] += downtime
            machines_data[machine_id]['weekly_data'][week_key]['idle'] += idle
            machines_data[machine_id]['weekly_data'][week_key]['machine_speed'] += speed
            machines_data[machine_id]['weekly_data'][week_key]['shift_count'] += 1
            
            machines_data[machine_id]['total_grade_a'] += grade_a
            machines_data[machine_id]['total_output'] += output
            machines_data[machine_id]['total_downtime'] += downtime
            machines_data[machine_id]['total_idle'] += idle
            machines_data[machine_id]['total_machine_speed'] += speed
            machines_data[machine_id]['shift_count'] += 1
        
        # Calculate monthly efficiency for each machine
        for machine_id in machines_data:
            data = machines_data[machine_id]
            
            # Calculate weekly efficiencies
            for week_key in data['weekly_data']:
                week_data = data['weekly_data'][week_key]
                speed_per_min = week_data['machine_speed'] / week_data['shift_count'] if week_data['shift_count'] > 0 else 0
                mrt = week_data['grade_a'] / speed_per_min if speed_per_min > 0 else 0
                total_time = mrt + week_data['downtime'] + week_data['idle']
                week_data['efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            
            # Calculate monthly average efficiency
            speed_per_min = data['total_machine_speed'] / data['shift_count'] if data['shift_count'] > 0 else 0
            mrt = data['total_grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + data['total_downtime'] + data['total_idle']
            data['avg_efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            data['mrt'] = round(mrt, 1)
            data['total_time'] = round(total_time, 1)
            data['quality'] = round((data['total_grade_a'] / data['total_output']) * 100, 1) if data['total_output'] > 0 else 0
        
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        # Month names in Indonesian
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        return jsonify({
            'year': year,
            'month': month,
            'month_name': month_names[month],
            'month_start': month_start.isoformat(),
            'month_end': month_end.isoformat(),
            'machines': result,
            'total_machines': len(result)
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
